from sqlalchemy import Column, Integer, String, create_engine
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, aliased, Query
import argparse
from openpyxl import load_workbook

engine = create_engine('postgresql+psycopg2://postgres:postgres@127.0.0.1:5435/postgres_period')
Base = declarative_base()
Session = sessionmaker(bind=engine)
session = Session()


class EndpointExel(Base):
    __tablename__ = 'endpoint'
    id = Column(Integer, primary_key=True)
    name = Column(String, nullable=True)

    def __init__(self, id, name):
        self.id = id
        self.name = name

    def __repr__ (self):
        return self.name


Base.metadata.create_all(engine)


parser = argparse.ArgumentParser(description='parser of path of exel_file')
parser.add_argument('exel_path')
args = parser.parse_args()
#print(args.exel_path)
workbook = load_workbook(filename=args.exel_path)
sheet = workbook.active
#print(sheet.max_column)
for values in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2, values_only=True):
    if values[0] != None:
        endpoint = EndpointExel(id=values[0], name=values[1])
        session.add(endpoint)
session.commit()

